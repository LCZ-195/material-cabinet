# -*- coding: utf-8 -*-
"""库存操作服务层（带日志）"""
from models.inventory_model import Inventory, Slot
from models.material_model import Material
from models.bom_model import OperationLog, BomRecord, BomItem

# 延迟初始化的同步服务实例（避免循环导入）
_sync_service = None


def _mark_inventory_changed():
    """通知 GitHubSyncService 库存发生变更，触发后台自动上传。"""
    global _sync_service
    try:
        if _sync_service is None:
            from services.github_sync_service import GitHubSyncService
            _sync_service = GitHubSyncService()
        _sync_service.mark_inventory_changed()
    except Exception:  # noqa: BLE001
        # 同步标记失败不能影响主业务流程
        pass


class InventoryService:
    """库存业务：入库、出库、调整"""

    @staticmethod
    def stock_in(slot_id, material_id, quantity, batch_no=None, note=None):
        """入库（单格多物料：同物料自动累加，不同物料作为新条目追加）"""
        if not material_id:
            raise ValueError("请先选择物料")
        if quantity <= 0:
            raise ValueError("数量必须大于0")
        inv_id = Inventory.add_inventory_to_slot(slot_id, material_id, quantity, batch_no, note)
        OperationLog.add(
            "inbound", "inventory", inv_id,
            {"slot_id": slot_id, "material_id": material_id,
             "quantity": quantity, "batch_no": batch_no, "note": note}
        )
        _mark_inventory_changed()
        return inv_id

    @staticmethod
    def stock_out(inv_id, quantity, note=None):
        """出库（从指定库存条目扣除，条件更新原子完成防并发超扣）"""
        try:
            quantity = int(quantity)
        except (TypeError, ValueError):
            raise ValueError("数量必须是整数")
        if quantity <= 0:
            raise ValueError("数量必须大于0")
        inv = Inventory.get(inv_id)
        if not inv:
            raise ValueError("库存记录不存在")
        if not Inventory.deduct_stock(inv_id, quantity):
            raise ValueError(
                f"库存不足（当前 {int(inv['quantity'])}，需取出 {quantity}）")
        OperationLog.add(
            "outbound", "inventory", inv_id,
            {"inventory_id": inv_id, "material_id": inv.get("material_id"),
             "slot_id": inv.get("slot_id"), "quantity": quantity, "note": note}
        )
        _mark_inventory_changed()
        return True

    @staticmethod
    def clear_slot(slot_id):
        """清空某格位的所有物料"""
        invs = Inventory.list_by_slot(slot_id)
        detail = {"slot_id": slot_id}
        if invs:
            detail["removed"] = [
                {"inv_id": i["id"], "material_id": i.get("material_id"),
                 "qty": i.get("quantity")} for i in invs
            ]
        OperationLog.add("clear", "slot", slot_id, detail)
        Inventory.remove_from_slot(slot_id)
        _mark_inventory_changed()

    @staticmethod
    def remove_single_inventory(slot_id, inv_id):
        """只移除格位中的某一条库存条目（保留该格其他物料）"""
        inv = Inventory.get(inv_id)
        OperationLog.add(
            "edit", "inventory", inv_id,
            {"action": "remove_from_slot", "slot_id": slot_id,
             "material_id": inv.get("material_id") if inv else None,
             "qty": inv.get("quantity") if inv else 0}
        )
        Inventory.remove_from_slot(slot_id, inv_id)
        _mark_inventory_changed()

    @staticmethod
    def adjust_stock(inv_id, new_qty, reason=""):
        """盘点调平（禁止负库存）"""
        try:
            new_qty = int(new_qty)
        except (TypeError, ValueError):
            raise ValueError("数量必须是整数")
        if new_qty < 0:
            raise ValueError("库存数量不能为负数")
        inv = Inventory.get(inv_id)
        if not inv:
            raise ValueError("库存记录不存在")
        old_qty = int(inv["quantity"])
        Inventory.update_quantity(inv_id, new_qty)
        OperationLog.add(
            "edit", "inventory", inv_id,
            {"inventory_id": inv_id, "old_qty": old_qty, "new_qty": new_qty,
             "reason": reason}
        )
        _mark_inventory_changed()

    @staticmethod
    def restock_from_bom(bom_id, auto_location=True):
        """执行补货BOM入库：按 suggested_slot_id 执行入库，写入日志+更新状态"""
        items = BomItem.list_by_bom(bom_id)
        ok = fail = skipped = 0
        used_slot_ids = set()  # 同一批补货已建议的格位，避免多行都推荐同一格
        for it in items:
            qty = int(it.get("required_qty") or 0) - int(it.get("picked_qty") or 0)
            if qty <= 0:
                skipped += 1
                continue
            mm = Material.find_or_create_from_bom(it)
            mid = mm["id"] if mm else None
            # 无法建档才跳过
            if not mid:
                fail += 1
                BomItem.update_match(it["id"], None, "unmatched", None,
                                      it.get("suggested_slot_id"))
                continue
            sid = it.get("suggested_slot_id")
            # auto location 没建议就智能推荐（排除已建议格位，顺序排布）
            if not sid and auto_location:
                suggs = Inventory.suggest_location_for_restock(
                    specification=it.get("specification"),
                    package=it.get("package"),
                    category=it.get("category") or it.get("mat_category"),
                    material_id=mid,
                    seed_key=it.get("material_code") or it.get("material_name"),
                    exclude_slot_ids=used_slot_ids,
                )
                if suggs:
                    sid = suggs[0]["slot_id"]
                    used_slot_ids.add(sid)
                    BomItem.update_match(it["id"], None, "restock_ok", None, sid)
            if not sid:
                fail += 1
                continue
            inv_id = InventoryService.stock_in(
                sid, mid, qty,
                batch_no=f"REPLENISH-{bom_id}",
                note=f"BOM#{bom_id} 行{it.get('line_no')} 补货入库"
            )
            BomItem.confirm_pick(it["id"], qty)
            BomItem.update_match(it["id"], inv_id, "restock_ok", None, sid)
            ok += 1
        OperationLog.add("restock", "bom", bom_id,
                         {"ok": ok, "fail": fail, "skipped": skipped})
        # restock_from_bom 内部已多次调用 stock_in，每次都标记了。
        # 这里不再重复标记，避免冗余。
        return ok, fail, skipped


class ExportService:
    """数据导出：库存清单、低库存、BOM领料单等"""

    @staticmethod
    def export_inventories(file_path: str) -> bool:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "库存清单"
        headers = ["格位编码", "物料编码", "物料名称", "分类", "规格", "封装",
                   "品牌", "立创编号", "供应商编号", "数量", "单位", "批次", "备注"]
        ws.append(headers)
        for inv in Inventory.all():
            ws.append([
                inv.get("slot_code", ""),
                inv.get("material_code", ""),
                inv.get("material_name", ""),
                inv.get("category", ""),
                inv.get("specification", ""),
                inv.get("package", ""),
                inv.get("brand", ""),
                inv.get("lcsc_code", ""),
                inv.get("supplier_code", ""),
                inv.get("quantity", 0),
                inv.get("unit", ""),
                inv.get("batch_no", ""),
                inv.get("note", ""),
            ])
        # 第二张：按格位汇总（一格多物料）
        ws2 = wb.create_sheet("按格位多物料汇总")
        ws2.append(["格位编码", "物料种数", "物料1名称/规格/数量",
                    "物料2名称/规格/数量", "物料3..."])
        from models.inventory_model import Slot as SlotModel
        for s in SlotModel.all():
            invs = s.get("all_inventories") or []
            row = [s["slot_code"], s.get("multi_count", 0)]
            for i in invs:
                row.append(f"{i.get('material_name','')} "
                           f"{i.get('specification','')} | {i.get('quantity',0)}"
                           f"{i.get('unit','个')} (品牌: {i.get('brand','')})")
            ws2.append(row)
        try:
            wb.save(file_path)
            return True
        except Exception:
            return False

    @staticmethod
    def export_materials(file_path: str) -> bool:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "物料主数据"
        headers = ["物料编码", "名称", "分类", "规格", "封装", "品牌",
                   "立创编号", "供应商编号", "单位", "最低库存", "描述", "数据手册"]
        ws.append(headers)
        for m in Material.all():
            ws.append([
                m.get("material_code", ""), m.get("name", ""), m.get("category", ""),
                m.get("specification", ""), m.get("package", ""), m.get("brand", ""),
                m.get("lcsc_code", ""), m.get("supplier_code", ""), m.get("unit", ""),
                m.get("min_stock", 0), m.get("description", ""), m.get("datasheet_url", ""),
            ])
        try:
            wb.save(file_path)
            return True
        except Exception:
            return False

    @staticmethod
    def export_bom_picklist(bom_id: int, file_path: str) -> bool:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        bom = BomRecord.get(bom_id)
        btype = bom.get("bom_type") if bom else "pick"
        ws.title = "补货单" if btype == "restock" else "领料单"
        if bom:
            ws.append(["BOM名称", bom.get("bom_name", "")])
            ws.append(["项目名称", bom.get("project_name", "")])
            ws.append(["类型", "补货入库" if btype == "restock" else "领料出库"])
            ws.append([])
        if btype == "restock":
            ws.append(["序号", "物料编码", "物料名称", "规格", "封装",
                       "需补数量", "已入数量", "建议存放格位", "格位内已有物料", "状态", "备注"])
        else:
            ws.append(["序号", "物料编码", "物料名称", "规格", "封装",
                       "需求数量", "已领数量", "所在格位", "当前库存", "状态", "备注"])
        status_map = {"unmatched": "未匹配", "partial": "部分",
                      "fully": "完成", "replaced": "替代", "restock_ok": "已入库"}
        for item in BomItem.list_by_bom(bom_id):
            st = status_map.get(item.get("match_status", ""), item.get("match_status", ""))
            if btype == "restock":
                from models.inventory_model import Inventory as Inv
                slot_code = item.get("slot_code") or ""
                # 若建议格位有，再补充内部已有物料
                existing = ""
                ssid = item.get("suggested_slot_id")
                if ssid:
                    invs = Inv.list_by_slot(ssid)
                    existing = "; ".join(
                        f"{i.get('material_name','')}x{i.get('quantity',0)}" for i in invs
                        if int(i.get("quantity") or 0) > 0
                    )
                    slot = Slot.get(ssid)
                    if slot:
                        slot_code = slot_code or slot.get("slot_code", "")
                ws.append([
                    item.get("line_no", ""), item.get("material_code", ""),
                    item.get("material_name", ""), item.get("specification", ""),
                    item.get("package", ""), item.get("required_qty", 0),
                    item.get("picked_qty", 0), slot_code,
                    existing, st, item.get("note", ""),
                ])
            else:
                ws.append([
                    item.get("line_no", ""), item.get("material_code", ""),
                    item.get("material_name", ""), item.get("specification", ""),
                    item.get("package", ""), item.get("required_qty", 0),
                    item.get("picked_qty", 0), item.get("slot_code", ""),
                    "", st, item.get("note", ""),
                ])
        try:
            wb.save(file_path)
            return True
        except Exception:
            return False
