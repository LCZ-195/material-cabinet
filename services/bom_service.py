# -*- coding: utf-8 -*-
"""BOM导入解析与比对服务"""
import os
import re
import logging
from typing import List, Dict, Tuple, Optional

from utils import PART_SEP_RE, is_lcsc_code

logger = logging.getLogger(__name__)


class BomImporter:
    """BOM文件解析器，支持CSV/Excel多种常见列名映射"""

    # 列名映射（支持多种命名习惯）
    COLUMN_ALIASES = {
        "material_code": ["物料编码", "物料编号", "料号", "PartNumber", "Part Number",
                          "PN", "Part No", "型号", "规格型号"],
        "material_name": ["物料名称", "名称", "品名", "Description", "器件名", "Part Name"],
        "specification": ["规格", "规格参数", "参数", "Spec", "Specification",
                          "Value", "值", "标称值"],
        "package": ["封装", "Package", "封装形式", "器件封装"],
        "footprint": ["Footprint", "Foot print", "PCB Footprint", "PCB封装", "PCB 封装"],
        "comment": ["Comment", "Comments", "注释", "说明注释", "器件描述"],
        "supplier_part": ["Supplier Part", "Supplier Part Number", "Supplier Part No",
                          "SupplierPart", "Supplier PartNumber", "供应商编号",
                          "供应商料号", "MPN", "Manufacturer Part Number",
                          "Manufacturer Part", "厂商编号", "厂商料号",
                          "LCSC Part #", "LCSC Part Number", "LCSC Part#",
                          "LCSC 编号", "立创编号"],
        "required_qty": ["数量", "需求数量", "Qty", "Quantity", "Order Qty",
                         "用量", "总数量", "Total Qty"],
        "line_no": ["序号", "行号", "No", "Index", "Item"],
        "category": ["分类", "Category", "Type", "类型"],
        "note": ["备注", "Note", "说明"],
    }

    @classmethod
    def parse_file(cls, file_path: str) -> Tuple[str, List[Dict]]:
        """解析BOM文件，返回 (错误信息, 解析出的行列表)

        每行至少需要: material_name 或 specification + required_qty
        """
        if not os.path.exists(file_path):
            return f"文件不存在: {file_path}", []
        ext = os.path.splitext(file_path)[1].lower()
        try:
            if ext == ".xls":
                return "暂不支持旧版 .xls 格式，请在 Excel 中另存为 .xlsx 后重新导入", []
            elif ext == ".xlsx":
                return "", cls._parse_excel(file_path)
            elif ext == ".csv":
                return "", cls._parse_csv(file_path)
            else:
                return f"不支持的文件格式: {ext}", []
        except Exception as e:
            logger.exception("BOM解析失败")
            return f"解析异常: {e}", []

    # 表头自动定位：最多扫描前几行（BOM 常见前几行为项目名/版本等说明）
    MAX_HEADER_SCAN_ROWS = 5

    @classmethod
    def _detect_header(cls, rows) -> Tuple[int, Dict[str, Optional[int]]]:
        """在前几行中自动定位表头行，返回 (表头行下标, 列映射)"""
        best_idx, best_mapping, best_hits = 0, None, -1
        for i, row in enumerate(rows[: cls.MAX_HEADER_SCAN_ROWS]):
            if row is None:
                continue
            header = [str(c).strip() if c is not None else "" for c in row]
            if not any(header):
                continue
            mapping = cls._match_columns(header)
            hits = sum(1 for v in mapping.values() if v is not None)
            has_key = any(
                mapping.get(k) is not None
                for k in ("material_code", "material_name", "specification")
            )
            if has_key and hits > best_hits:
                best_idx, best_mapping, best_hits = i, mapping, hits
        if best_mapping is None:
            # 兜底：取第一个非空行作为表头（防止首行为 None 时崩溃）
            for row in rows:
                if row is None:
                    continue
                cells = [str(c).strip() if c is not None else "" for c in row]
                if any(cells):
                    best_mapping = cls._match_columns(cells)
                    break
            else:
                best_mapping = cls._match_columns([])
        return best_idx, best_mapping

    @classmethod
    def _parse_excel(cls, path: str) -> List[Dict]:
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True)

        # ponytail: 旧逻辑只读 wb.active，彩色封面/"说明"页常被设为活动表，
        # 导致真实数据表解析为 0 或垃圾行；改为遍历全部工作表，
        # 选表头命中数最高（其次数据行最多）的一张。
        best = None  # (hits, data_rows, sheet_rows, header_idx, mapping)
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header_idx, mapping = cls._detect_header(rows)
            hits = sum(1 for v in mapping.values() if v is not None)
            data_rows = sum(
                1 for row in rows[header_idx + 1:]
                if row is not None and any(
                    c is not None and str(c).strip() != "" for c in row)
            )
            score = (hits, data_rows)
            if best is None or score > best[0]:
                best = (score, rows, header_idx, mapping)
        if best is None:
            return []
        _, rows, header_idx, mapping = best

        result = []
        for idx, row in enumerate(rows[header_idx + 1:], start=header_idx + 1):
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            item = {"line_no": idx}
            for field, col_idx in mapping.items():
                if col_idx is not None and col_idx < len(row):
                    val = row[col_idx]
                    if val is not None:
                        item[field] = str(val).strip() if not isinstance(val, (int, float)) else val
            # 数量转整数
            if "required_qty" in item:
                try:
                    item["required_qty"] = int(float(str(item["required_qty"]).replace(",", "")))
                except (ValueError, TypeError):
                    item["required_qty"] = 0
            else:
                item["required_qty"] = 0
            result.append(item)
        return result

    @staticmethod
    def _read_csv_rows(path: str) -> List[List[str]]:
        """读取 CSV，编码自动回退（兼容 Excel 导出的 GBK 编码）"""
        import csv
        for enc in ("utf-8-sig", "gbk", "gb18030"):
            try:
                with open(path, "r", encoding=enc, newline="") as f:
                    return list(csv.reader(f))
            except UnicodeDecodeError:
                continue
        with open(path, "r", encoding="utf-8", errors="replace", newline="") as f:
            return list(csv.reader(f))

    @classmethod
    def _parse_csv(cls, path: str) -> List[Dict]:
        rows = cls._read_csv_rows(path)
        if not rows:
            return []
        header_idx, mapping = cls._detect_header(rows)
        result = []
        for idx, row in enumerate(rows[header_idx + 1:], start=header_idx + 1):
            if not row or all(c.strip() == "" for c in row):
                continue
            item = {"line_no": idx}
            for field, col_idx in mapping.items():
                if col_idx is not None and col_idx < len(row):
                    item[field] = row[col_idx].strip()
            if "required_qty" in item:
                try:
                    item["required_qty"] = int(float(item["required_qty"].replace(",", "")))
                except (ValueError, TypeError):
                    item["required_qty"] = 0
            else:
                item["required_qty"] = 0
            result.append(item)
        return result

    @classmethod
    def _match_columns(cls, header: List[str]) -> Dict[str, Optional[int]]:
        """根据表头匹配字段位置：先精确、后子串，且一列最多归一个字段。

        ponytail: 旧逻辑的子串匹配会把"规格"列同时误映射为 material_code
        （"规格" ⊂ "规格型号"），导致同规格、不同供应商编号的两行被当成
        同编码物料错误合并（73 行 BOM 只剩 71 行），故改为精确命中优先 +
        列独占两轮匹配。
        """
        result = {k: None for k in cls.COLUMN_ALIASES}
        used = set()
        header_l = [h.strip().lower() if h else "" for h in header]
        # 第一轮：精确匹配（按字段、别名顺序，先到先得）
        for field, aliases in cls.COLUMN_ALIASES.items():
            for alias in aliases:
                alias_l = alias.lower()
                for i, h in enumerate(header_l):
                    if not h or i in used:
                        continue
                    if alias_l == h:
                        result[field] = i
                        used.add(i)
                        break
                if result[field] is not None:
                    break
        # 第二轮：子串匹配（较短方至少 2 字符，防止单字误命中）
        for field, aliases in cls.COLUMN_ALIASES.items():
            if result[field] is not None:
                continue
            for alias in aliases:
                alias_l = alias.lower()
                for i, h in enumerate(header_l):
                    if not h or i in used:
                        continue
                    if min(len(alias_l), len(h)) < 2:
                        continue
                    if alias_l in h or h in alias_l:
                        result[field] = i
                        used.add(i)
                        break
                if result[field] is not None:
                    break
        return result

    # 供应商编号分隔符与立创编号判断统一使用 utils（与库存/物料侧同一口径）
    _LCSC_RE = re.compile(r"c\d{4,}")

    @classmethod
    def part_keys(cls, item: Dict) -> set:
        """行内供应商编号集合（小写归一化）；comment 中的立创编号一并纳入，
        兼容多备选编号写在 Comment 列的 BOM 导出习惯"""
        keys = {p.strip().lower()
                for p in PART_SEP_RE.split(str(item.get("supplier_part") or ""))
                if p.strip()}
        for token in PART_SEP_RE.split(str(item.get("comment") or "")):
            token = token.strip().lower()
            if cls._LCSC_RE.fullmatch(token):
                keys.add(token)
        return keys

    @classmethod
    def same_material(cls, a: Dict, b: Dict) -> bool:
        """两行是否同一物料（导入合并的统一判定）：
        ① 内部编码相同且非空；② 供应商编号集合有交集（支持一行多个备选编号）；
        ③ 均无编号时，参数（名称+规格+封装）一致视为同物料。"""
        ca = str(a.get("material_code") or "").strip().lower()
        cb = str(b.get("material_code") or "").strip().lower()
        if ca and cb:
            return ca == cb
        ka, kb = cls.part_keys(a), cls.part_keys(b)
        if ka or kb:
            return bool(ka & kb)
        pk = ("material_name", "specification", "package")
        return (tuple(str(a.get(k) or "").strip().lower() for k in pk)
                == tuple(str(b.get(k) or "").strip().lower() for k in pk))

    @classmethod
    def union_parts(cls, items: List[Dict], field: str) -> str:
        """多行的编号/备注取并集（排序保证幂等），供行合并与追加入库复用"""
        seen, vals = set(), []
        for x in items:
            for piece in PART_SEP_RE.split(str(x.get(field) or "")):
                piece = piece.strip()
                if piece and piece.lower() not in seen:
                    seen.add(piece.lower())
                    vals.append(piece)
        return ",".join(sorted(vals, key=str.lower))

    @classmethod
    def merge_rows(cls, rows: List[Dict]) -> List[Dict]:
        """合并同物料行（同一文件内重复行 / 多文件追加导入），数量累加、
        编号与备注取并集、其余字段取首个非空值。

        ponytail: 传递合并由"任一成员命中即入组"自然达成——
        如多备选编号行(C14858,C277507)与逐编号行 C14858、C277507 归为一组。
        """
        groups: List[Dict] = []
        for it in rows:
            hits = [g for g in groups
                    if any(cls.same_material(m, it) for m in g["members"])]
            if not hits:
                groups.append({"members": [it]})
                continue
            target = hits[0]
            for g in hits[1:]:
                target["members"].extend(g["members"])
                groups = [h for h in groups if h is not g]
            target["members"].append(it)
        merged = []
        for g in groups:
            items = g["members"]
            if len(items) == 1:
                merged.append(items[0])
                continue
            base = dict(items[0])
            base["required_qty"] = sum(int(x.get("required_qty") or 0) for x in items)
            for field in ("supplier_part", "comment"):
                base[field] = cls.union_parts(items, field)
            for field in ("material_code", "material_name", "specification",
                          "package", "footprint", "note"):
                if not str(base.get(field) or "").strip():
                    base[field] = next((str(x.get(field) or "").strip()
                                        for x in items[1:]
                                        if str(x.get(field) or "").strip()), "")
            merged.append(base)
        return merged


class BomMatcher:
    """BOM比对：将BOM行与库存匹配"""

    @staticmethod
    def match_bom_item(item: Dict) -> Tuple[str, Optional[int], Optional[List[Dict]]]:
        """比对单条BOM行（融合匹配，多源并发）

        返回: (状态, 匹配的库存ID, 候选替代物料列表)
            状态: unmatched / partial / fully / replaced
        融合匹配优先级（①-③ 并发执行以缩短匹配时间）：
            ① 联网立创商城：供应商编号（厂商型号/立创编号）精确匹配
            ② 参数匹配：本地库存按 供应商编号→编码→规格+封装→名称 逐级匹配
            ③ 本地记忆库（内置元件库）：编号→关键词→规格封装
            ④ ①-③ 均未命中：立创 API 关键词搜索兜底，命中自动沉淀记忆库
        """
        from models.inventory_model import Inventory
        from models.component_lib_model import ComponentLib
        from concurrent.futures import ThreadPoolExecutor

        code = item.get("material_code") or ""
        spec = item.get("specification") or ""
        pkg = item.get("package") or ""
        name = item.get("material_name") or ""
        sp = item.get("supplier_part") or ""
        req_qty = int(item.get("required_qty") or 0)
        if req_qty <= 0:
            return "unmatched", None, []

        def _lookup(model=None, specification=None, package=None, mname=None,
                    msup=None, mscode=None):
            return Inventory.search_for_bom(
                material_code=model if model is not None else code,
                specification=specification if specification is not None else spec,
                package=package if package is not None else pkg,
                name=mname if mname is not None else name,
                supplier_part=msup if msup is not None else sp,
                **({"lcsc_code": mscode} if mscode else {}),
            )

        # ---------- 并发匹配源（①-③ 同时启动） ----------
        def _match_lcsc_sp():
            """① 立创供应商编号匹配（联网，仅供应商编号相关，命中沉淀记忆库）"""
            try:
                resolved = BomMatcher._resolve_via_lcsc_sp(code, sp, name, spec, pkg)
                if not resolved:
                    return []
                lcsc_new = resolved.get("lcsc_code") or \
                    (code.upper() if is_lcsc_code(code) else "")
                if lcsc_new or (resolved.get("supplier_part") or sp) or resolved.get("model"):
                    ComponentLib.record({
                        "lcsc_code": lcsc_new,
                        "supplier_part": resolved.get("supplier_part") or sp,
                        "model": resolved.get("model") or "",
                        "name": resolved.get("name") or "",
                        "specification": resolved.get("specification") or "",
                        "package": resolved.get("package") or "",
                        "brand": resolved.get("brand") or "",
                        "parameters": resolved.get("parameters") or None,
                    }, source="lcsc-search")
                return _lookup(
                    model=resolved.get("model") or code,
                    specification=resolved.get("specification") or spec,
                    package=resolved.get("package") or pkg,
                    mname=resolved.get("name") or name,
                    msup=resolved.get("supplier_part") or sp,
                )
            except Exception:
                logger.exception("① 立创供应商编号匹配失败")
                return []

        def _match_local():
            """② 参数匹配（本地库存：供应商编号→编码→规格+封装→名称）"""
            try:
                return _lookup()
            except Exception:
                logger.exception("② 本地参数匹配失败")
                return []

        def _match_lib():
            """③ 本地记忆库（内置元件库，离线秒回）"""
            try:
                resolved = BomMatcher._resolve_via_lib(code, sp, spec, pkg, name)
                if not resolved:
                    return []
                return _lookup(
                    model=resolved.get("model") or code,
                    specification=resolved.get("specification") or spec,
                    package=resolved.get("package") or pkg,
                    mname=resolved.get("name") or name,
                    msup=resolved.get("supplier_part") or sp,
                    mscode=resolved.get("lcsc_code") or None,
                )
            except Exception:
                logger.exception("③ 本地记忆库匹配失败")
                return []

        matches = []
        try:
            with ThreadPoolExecutor(max_workers=3) as ex:
                f_sp = ex.submit(_match_lcsc_sp)
                f_local = ex.submit(_match_local)
                f_lib = ex.submit(_match_lib)
                # 本地两路秒回：命中立即返回，避免等慢速联网卡住（立创作补充）
                for fut in (f_local, f_lib):
                    try:
                        rows = fut.result(timeout=5)
                    except Exception:  # noqa: BLE001
                        rows = []
                    if rows:
                        matches = rows
                        break
                if not matches:
                    # 本地未命中 → 立创供应商编号匹配限时等待（防止卡死）
                    try:
                        rows = f_sp.result(timeout=8)
                    except Exception:  # noqa: BLE001
                        rows = []
                    if rows:
                        matches = rows
        except Exception:  # noqa: BLE001
            logger.exception("融合匹配并发执行失败，回退本地匹配")
            matches = _match_local() or []

        if not matches:
            # ④ API 关键词兜底（含沉淀，限时 8s 防卡）。
            # 有 C 编号/供应商编号的行已在 ① 精确处理（含沉淀），跳过避免重复联网
            resolved = None
            if not (code or "").strip() and not (sp or "").strip():
                try:
                    with ThreadPoolExecutor(max_workers=1) as ex:
                        resolved = ex.submit(
                            BomMatcher._resolve_via_lcsc, code, spec, pkg, name, sp
                        ).result(timeout=8)
                except Exception:  # noqa: BLE001
                    resolved = None
            if resolved:
                try:
                    lcsc_new = resolved.get("lcsc_code") or \
                        (code.upper() if is_lcsc_code(code) else "")
                    if lcsc_new or (resolved.get("supplier_part") or sp) or resolved.get("model"):
                        ComponentLib.record({
                            "lcsc_code": lcsc_new,
                            "supplier_part": resolved.get("supplier_part") or sp,
                            "model": resolved.get("model") or "",
                            "name": resolved.get("name") or "",
                            "specification": resolved.get("specification") or "",
                            "package": resolved.get("package") or "",
                            "brand": resolved.get("brand") or "",
                            "parameters": resolved.get("parameters") or None,
                        }, source="lcsc-search")
                except Exception:  # noqa: BLE001
                    logger.exception("内置元件库沉淀失败（不影响匹配）")
                matches = _lookup(
                    model=resolved.get("model") or code,
                    specification=resolved.get("specification") or spec,
                    package=resolved.get("package") or pkg,
                    mname=resolved.get("name") or name,
                    msup=resolved.get("supplier_part") or sp,
                )
        if not matches:
            # 全部未命中：找规格相近的替代物料
            return "unmatched", None, BomMatcher._find_replacement_candidates(item)

        # 找一个数量足够的
        chosen = None
        for m in matches:
            if int(m.get("quantity", 0)) >= req_qty:
                chosen = m
                break
        if chosen is None:
            chosen = matches[0]  # 数量不够但至少匹配到
            status = "partial"
        else:
            status = "fully"
        return status, chosen["id"], None

    @staticmethod
    def _resolve_via_lcsc_sp(code: str, supplier_part: str, name: str,
                             spec: str = "", pkg: str = "") -> Optional[Dict]:
        """① 联网立创商城供应商编号精确匹配。

        仅处理有供应商编号/立创编号的行：C 编号走详情查询，厂商型号走关键词
        搜索取最相似首条；没有任何编号时直接返回 None，避免无谓的联网请求。
        返回 {name, model, specification, package, brand, lcsc_code,
              supplier_part, parameters}，离线/未命中返回 None。"""
        from services.lcsc_service import LCSCApi

        api = LCSCApi.get_shared()
        code = (code or "").strip()
        sp = (supplier_part or "").strip()
        detail = None
        if is_lcsc_code(code):
            detail = api.get_product_detail(code.upper())
        elif is_lcsc_code(sp):
            # Supplier Part 列直接填了立创编号时也走详情查询
            detail = api.get_product_detail(sp.upper())
        elif sp:
            # 供应商编号（厂商型号）关键词搜索，取最相似首条
            rows = api.search_product(sp, page_size=5) or []
            if rows:
                r0 = rows[0]
                detail = {
                    "name": r0.get("model") or name,
                    "model": r0.get("model") or sp,
                    "specification": r0.get("specification") or spec,
                    "package": r0.get("package") or pkg,
                    "brand": r0.get("brand") or "",
                    "lcsc_code": r0.get("lcsc_code") or "",
                    "supplier_part": sp,
                    "parameters": r0.get("parameters") or {},
                }
        else:
            return None
        if isinstance(detail, dict) and detail:
            if "lcsc_code" not in detail:
                detail["lcsc_code"] = code.upper() or ""
            return detail
        return None

    @staticmethod
    def _resolve_via_lib(code: str, supplier_part: str, spec: str,
                         pkg: str, name: str) -> Optional[Dict]:
        """内置元件库解析：编号精确 → 关键词 → 规格封装。命中自动累计命中数。"""
        from models.component_lib_model import ComponentLib

        rec = None
        lcsc = code.upper() if is_lcsc_code(code) else ""
        if supplier_part or lcsc:
            rec = ComponentLib.find_by_code(supplier_part=supplier_part or None,
                                            lcsc_code=lcsc or None)
        if not rec:
            kw = supplier_part or code or name or spec
            hits = ComponentLib.search(kw, limit=1) if kw else []
            rec = hits[0] if hits else None
        if not rec:
            hits = ComponentLib.search_by_spec(spec or None, pkg or None, limit=1)
            rec = hits[0] if hits else None
        if not rec:
            return None
        try:
            ComponentLib.bump_hit(rec["id"])
        except Exception:
            logger.exception("元件库命中计数失败")
        return {
            "name": rec.get("name") or rec.get("model") or name,
            "model": rec.get("model") or rec.get("supplier_part") or code,
            "specification": rec.get("specification") or spec,
            "package": rec.get("package") or pkg,
            "supplier_part": rec.get("supplier_part") or supplier_part,
            "lcsc_code": rec.get("lcsc_code") or lcsc,
            "brand": rec.get("brand") or "",
        }

    @staticmethod
    def _resolve_via_lcsc(code: str, spec: str, pkg: str, name: str,
                          supplier_part: str = "") -> Optional[Dict]:
        """联网立创商城解析 BOM 行（C编号/供应商编号优先查详情，否则关键词搜索）。
        返回 {name, model, specification, package, brand, lcsc_code,
              supplier_part, parameters}，离线/未命中返回 None。"""
        from services.lcsc_service import LCSCApi

        api = LCSCApi.get_shared()
        code = (code or "").strip()
        sp = (supplier_part or "").strip()
        detail = None
        if is_lcsc_code(code):
            detail = api.get_product_detail(code.upper())
        elif is_lcsc_code(sp):
            # Supplier Part 列直接填了立创编号时也走详情查询
            detail = api.get_product_detail(sp.upper())
        elif sp:
            # 供应商编号（厂商型号）关键词搜索，取最相似首条
            rows = api.search_product(sp, page_size=5) or []
            if rows:
                r0 = rows[0]
                detail = {
                    "name": r0.get("model") or name,
                    "model": r0.get("model") or sp,
                    "specification": r0.get("specification") or spec,
                    "package": r0.get("package") or pkg,
                    "brand": r0.get("brand") or "",
                    "lcsc_code": r0.get("lcsc_code") or "",
                    "supplier_part": sp,
                    "parameters": r0.get("parameters") or {},
                }
        if isinstance(detail, dict) and detail and "lcsc_code" not in detail:
            intro = detail.get("specification") or detail.get("name") or ""
            detail = {
                "name": intro or name,
                "model": detail.get("model") or code,
                "specification": intro or spec,
                "package": detail.get("package") or pkg,
                "brand": detail.get("brand") or "",
                "lcsc_code": code.upper(),
                "supplier_part": detail.get("model") or sp,
                "parameters": detail.get("parameters") or {},
            }
        if detail:
            return detail
        keyword = sp or name or spec or code
        if not keyword:
            return None
        rows = api.search_product(keyword, page_size=5) or []
        keys = [str(k).lower() for k in (spec, pkg) if k]
        for r in rows:
            hay = f"{r.get('model', '')} {r.get('specification', '')} {r.get('package', '')}".lower()
            if all(k in hay for k in keys):
                return {
                    "name": r.get("model") or name,
                    "model": r.get("model") or code,
                    "specification": r.get("specification") or spec,
                    "package": r.get("package") or pkg,
                    "brand": r.get("brand") or "",
                    "lcsc_code": r.get("lcsc_code") or "",
                    "supplier_part": sp or r.get("model") or "",
                    "parameters": r.get("parameters") or {},
                }
        return None

    @staticmethod
    def _find_replacement_candidates(item: Dict) -> List[Dict]:
        from models.material_model import Material
        from services.lcsc_service import LocalParameterMatcher
        # 构造一个假的 material 对象走本地匹配器
        pseudo_mat = {
            "id": None,
            "specification": item.get("specification", ""),
            "package": item.get("package", ""),
            "category": item.get("category", ""),
            "name": item.get("material_name", ""),
        }
        matcher = LocalParameterMatcher()
        candidates = matcher.find_candidates_from_db(pseudo_mat)
        # 只保留有库存的
        with_inv = []
        for c in candidates:
            from models.inventory_model import Inventory
            invs = Inventory.get_by_material(c["id"])
            if invs:
                c["_inventories"] = invs
                with_inv.append(c)
        return with_inv[:5]
